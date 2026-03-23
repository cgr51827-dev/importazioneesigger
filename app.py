import streamlit as st
import pandas as pd
from io import BytesIO
import zipfile
from datetime import datetime
from openpyxl import load_workbook

st.set_page_config(page_title="GeCO Generator", layout="wide")

# -----------------------
# LOGIN
# -----------------------
def login():
    if "logged" not in st.session_state:
        st.session_state.logged = False

    if not st.session_state.logged:
        st.title("🔐 Accesso")
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")

        if st.button("Accedi"):
            if username == "RECAP" and password == "Recap26@":
                st.session_state.logged = True
                st.rerun()
            else:
                st.error("Credenziali errate")

        st.stop()

login()

st.title("📄 GeCO File Generator")
st.markdown("Carica i file e genera automaticamente Import Standard e Recapiti")

# -----------------------
# MAPPA CSV
# -----------------------
CSV_MAP = {
    "A": "POS_NUM",
    "B": "POS_COD",
    "C": "LOTTO",
    "D": "DATA_AFFIDAMENTO",
    "E": "DATA_SCADENZA",
    "F": "CAPITALE",
    "G": "INTERESSI",
    "H": "ONERI",
    "I": "TOTALE",
    "J": "DBT_RAGIONESOCIALE",
    "K": "DBT_INDIRIZZO",
    "L": "DBT_CAP",
    "M": "DBT_COMUNE",
    "N": "DBT_PROVINCIA",
    "O": "DBT_CODFISCALE",
    "P": "DBT_PIVA",
    "Q": "TEL1",
    "R": "TEL2",
    "S": "TEL3",
    "T": "TEL4",
    "U": "TEL5",
    "V": "TEL6",
    "W": "EMAIL",
    "X": "NOTE1",
    "Y": "NOTE2",
}

REQUIRED_REAL_COLUMNS = [
    "POS_NUM",
    "CAPITALE",
    "INTERESSI",
    "ONERI",
    "DBT_RAGIONESOCIALE",
    "DBT_INDIRIZZO",
    "DBT_CAP",
    "DBT_COMUNE",
    "DBT_PROVINCIA",
    "DBT_CODFISCALE",
    "TEL1",
    "TEL2",
    "TEL3",
    "TEL4",
    "TEL5",
    "TEL6",
    "EMAIL",
]

# -----------------------
# FUNZIONI
# -----------------------
def normalize_columns(df):
    df.columns = [str(c).strip().upper() for c in df.columns]
    return df

def read_csv_robust(uploaded_file):
    raw = uploaded_file.getvalue()

    encodings = ["utf-8-sig", "utf-8", "latin1", "cp1252"]
    seps = [";", ",", "\t", "|"]

    last_error = None

    for enc in encodings:
        for sep in seps:
            try:
                df = pd.read_csv(
                    BytesIO(raw),
                    sep=sep,
                    encoding=enc,
                    dtype=str,
                    keep_default_na=False,
                )
                df = normalize_columns(df)
                if len(df.columns) >= 10:
                    return df
            except Exception as e:
                last_error = e

    raise last_error if last_error else Exception("Impossibile leggere il CSV")

def check_required_columns(df):
    return [c for c in REQUIRED_REAL_COLUMNS if c not in df.columns]

def get_csv_value(row, logical_col):
    real_col = CSV_MAP.get(logical_col, "")
    return row.get(real_col, "")

def add_zero_if_needed(value):
    if pd.isna(value):
        return ""

    value = str(value).strip()

    if value == "" or value.lower() == "nan":
        return ""

    if value.startswith("+"):
        return value

    if value.startswith("00"):
        return value

    if value.startswith("0"):
        return value

    digits = "".join(ch for ch in value if ch.isdigit())
    if digits == "":
        return value

    return "0" + digits

def clear_sheet_data(ws, start_row=2, max_col=22):
    if ws.max_row >= start_row:
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.value = None

def build_import_workbook(template_file, df):
    wb = load_workbook(template_file)
    ws = wb.active

    clear_sheet_data(ws, start_row=2, max_col=22)

    excel_row = 2
    for _, row in df.iterrows():
        ws.cell(excel_row, 1).value = get_csv_value(row, "J")                         # A <- J
        ws.cell(excel_row, 4).value = get_csv_value(row, "O")                         # D <- O
        ws.cell(excel_row, 5).value = ""                                              # E vuota
        ws.cell(excel_row, 9).value = get_csv_value(row, "W")                         # I <- W
        ws.cell(excel_row, 10).value = get_csv_value(row, "K")                        # J <- K
        ws.cell(excel_row, 11).value = get_csv_value(row, "L")                        # K <- L
        ws.cell(excel_row, 12).value = get_csv_value(row, "M")                        # L <- M
        ws.cell(excel_row, 13).value = get_csv_value(row, "N")                        # M <- N
        ws.cell(excel_row, 14).value = get_csv_value(row, "F")                        # N <- F
        ws.cell(excel_row, 15).value = get_csv_value(row, "H")                        # O <- H
        ws.cell(excel_row, 16).value = get_csv_value(row, "G")                        # P <- G
        ws.cell(excel_row, 21).value = add_zero_if_needed(get_csv_value(row, "A"))    # U <- A con 0
        excel_row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def build_recap_workbook(template_file, df):
    wb = load_workbook(template_file)
    ws = wb.active

    clear_sheet_data(ws, start_row=2, max_col=22)

    excel_row = 2
    for _, row in df.iterrows():
        ws.cell(excel_row, 2).value = add_zero_if_needed(get_csv_value(row, "A"))     # B <- A con 0

        telefoni = []
        for logical_col in ["Q", "R", "S", "T", "U", "V"]:
            val = get_csv_value(row, logical_col)
            if str(val).strip() != "":
                telefoni.append(add_zero_if_needed(val))

        # H..V = colonne 8..22
        target_col = 8
        for tel in telefoni:
            if target_col <= 22:
                ws.cell(excel_row, target_col).value = tel
                target_col += 1

        excel_row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# -----------------------
# UPLOAD
# -----------------------
file_csv = st.file_uploader("📂 File madre (CSV)", type=["csv"])
file_import = st.file_uploader("📂 Template Import Standard (.xlsx)", type=["xlsx"])
file_recap = st.file_uploader("📂 Template Recapiti (.xlsx)", type=["xlsx"])

# -----------------------
# GENERAZIONE
# -----------------------
if st.button("🚀 Genera File"):
    if not file_csv or not file_import or not file_recap:
        st.error("❌ Carica tutti i file")
        st.stop()

    try:
        df = read_csv_robust(file_csv)
    except Exception:
        st.error("❌ Impossibile leggere il CSV. Controlla formato, separatore o encoding.")
        st.stop()

    missing = check_required_columns(df)
    if missing:
        st.error(f"❌ Colonne mancanti nel CSV: {missing}")
        st.write("Colonne trovate:", list(df.columns))
        st.stop()

    try:
        import_buffer = build_import_workbook(file_import, df)
        file_import.seek(0)
        recap_buffer = build_recap_workbook(file_recap, df)
    except Exception as e:
        st.error(f"❌ Errore nella compilazione dei template: {e}")
        st.stop()

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("GeCO_import_standard.xlsx", import_buffer.getvalue())
        z.writestr("GeCO_recapiti.xlsx", recap_buffer.getvalue())

    zip_buffer.seek(0)

    st.success(f"✅ File generati correttamente. Pratiche elaborate: {len(df)}")

    st.download_button(
        "⬇️ Scarica ZIP",
        data=zip_buffer,
        file_name=f"geco_output_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
        mime="application/zip",
    )
